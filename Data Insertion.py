import mysql.connector
import openpyxl

mydb = mysql.connector.connect(host="localhost", user="root", password="2510triblaz", db="Car_Rental_Service")
db_cursor = mydb.cursor()

mainlist = []

Open_file = openpyxl.load_workbook("Resources/Car_details.xlsx")

sheet = Open_file.active

max_col = sheet.max_column
max_row = sheet.max_row


for i in range(2, max_row+1):
    list = []
    for j in range(1, max_col+1):
        cell_obj = sheet.cell(row=i, column=j)
        list.append(cell_obj.value)
    mainlist.append(tuple(list))
print(list)

Query = "insert into specifications(Car_name, type, colour, transmission, fuel, mileage, seats" \
                  ", ac, infotainment, foglamps) values(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
db_cursor.executemany(Query, mainlist)
mydb.commit()
mydb.close()
